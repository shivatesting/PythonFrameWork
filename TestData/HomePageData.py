import os

import openpyxl
import pytest


class HomePageData:

    @staticmethod
    def getTestData(test_case_name):
        Dict = {}  # Initialize an empty dictionary to store test data
        current_directory = os.path.dirname(os.path.realpath(__file__))  # Get the current directory where the script is located
        file_path = os.path.join(current_directory, "PythonDemo.xlsx")  # Construct the file path to the Excel file containing test data
        book = openpyxl.load_workbook(file_path) # Open the Excel workbook
        sheet = book.active # Get the active sheet in the workbook

        # Iterate over each row in the Excel sheet
        for i in range(1, sheet.max_row + 1):
            # Check if the value in the first column of the current row matches the given test case name
            if sheet.cell(row=i, column=1).value == test_case_name:
                # If there's a match, iterate over each column in the row (starting from the second column)
                for j in range(2, sheet.max_column + 1):
                    # Assign the value of each cell in the row (except the first column)
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [Dict]  # Return a list containing the dictionary with test data


